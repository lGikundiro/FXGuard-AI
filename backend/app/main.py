from __future__ import annotations

import json
import mimetypes
import os
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from backend.app.accounts import configured_origins, router as accounts_router
from backend.app.rates import (
    CURRENCY_INFO,
    SUPPORTED_CURRENCIES,
    combined_daily,
    latest_feature_row as latest_currency_feature_row,
    provider_status,
    validate_currency,
)

ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data" / "processed"
MODEL_DIR = ROOT / "backend" / "models"
FRONTEND_DIR = ROOT / "frontend"
FEEDBACK_DIR = ROOT / "reports" / "feedback"
FEEDBACK_FILE = FEEDBACK_DIR / "prototype_feedback.xlsx"
FEEDBACK_FORM_NAME = os.getenv("FEEDBACK_FORM_NAME", "FXGuard AI User feedback")
GOOGLE_SHEETS_WEBHOOK_URL = os.getenv("GOOGLE_SHEETS_WEBHOOK_URL", "").strip()
GOOGLE_FORM_RESPONSE_URL = os.getenv(
    "GOOGLE_FORM_RESPONSE_URL",
    (
        "https://docs.google.com/forms/d/e/"
        "1FAIpQLSd3E97VFGFl7v-9ojSAAmPc4RkE-30tf9YCJ_XUhPuw8JFbBg/formResponse"
    ),
).strip()
GOOGLE_FORM_ENTRY_IDS = {
    "clarity_rating": "entry.1987006101",
    "usefulness_rating": "entry.1258240246",
    "comment": "entry.770935896",
    "participant_name": "entry.792416587",
    "import_category": "entry.2052407924",
    "phone_number": "entry.1877013245",
}
GOOGLE_FORM_IMPORT_CATEGORIES = {
    "Raw materials",
    "Minerals",
    "Finished goods",
    "Medicine",
}

FEATURE_COLUMNS = [
    "mid_rate", "daily_return", "return_7d", "return_14d", "ma_7", "ma_14", "ma_30",
    "ma_gap", "volatility_7d", "volatility_14d", "volatility_30d", "momentum_7d",
    "momentum_14d", "spread", "spread_pct", "depreciation_days_7d", "depreciation_days_14d",
]
MODEL_FEATURE_COLUMNS = [*FEATURE_COLUMNS, "horizon_days"]
MIN_PAYMENT_DAYS = 1
MAX_PAYMENT_DAYS = 100
RWANDA_TIMEZONE = timezone(timedelta(hours=2))


def rwanda_today() -> date:
    return datetime.now(RWANDA_TIMEZONE).date()

app = FastAPI(
    title="FXGuard AI API",
    description="Multi-currency exchange-rate risk forecasting and decision support for Rwanda-based importers.",
    version="2.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=configured_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.include_router(accounts_router)

mimetypes.add_type("font/woff2", ".woff2")

if FRONTEND_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")


class RiskRequest(BaseModel):
    currency: str = Field(default="USD", description="Invoice currency: USD, EUR, or KES.")
    amount: float = Field(default=10000, gt=0, description="Supplier invoice amount in the selected currency.")
    payment_date: date = Field(description="Planned invoice payment date, 1 to 100 days from today.")


class FeedbackRequest(BaseModel):
    participant_name: Optional[str] = Field(default=None, max_length=100)
    import_category: Optional[str] = Field(default=None, max_length=100)
    phone_number: Optional[str] = Field(default=None, max_length=30)
    clarity_rating: int = Field(ge=1, le=5)
    usefulness_rating: int = Field(ge=1, le=5)
    comment: str = Field(min_length=1, max_length=2000)


FEEDBACK_COLUMNS = [
    "submitted_at",
    "participant_name",
    "import_category",
    "phone_number",
    "clarity_rating",
    "usefulness_rating",
    "comment",
]


_cache = {}


def load_json(path: Path):
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def project_path_label(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def append_feedback_to_google_sheet(feedback_data: dict) -> dict:
    if GOOGLE_SHEETS_WEBHOOK_URL:
        payload = {
            "form_name": FEEDBACK_FORM_NAME,
            "columns": FEEDBACK_COLUMNS,
            "response": {column: feedback_data.get(column) for column in FEEDBACK_COLUMNS},
        }
        data = json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            GOOGLE_SHEETS_WEBHOOK_URL,
            data=data,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(request, timeout=8) as response:
            response_text = response.read().decode("utf-8").strip()
            response_body = json.loads(response_text) if response_text else {}
        if isinstance(response_body, dict) and response_body.get("status") in {"saved", "ok"}:
            return {"enabled": True, "status": "saved", "method": "webhook"}
        return {"enabled": True, "status": "unknown_response", "method": "webhook"}

    if not GOOGLE_FORM_RESPONSE_URL:
        return {"enabled": False, "status": "not_configured"}

    form_values = {
        GOOGLE_FORM_ENTRY_IDS[field]: str(feedback_data.get(field) or "")
        for field in GOOGLE_FORM_ENTRY_IDS
        if field != "import_category"
    }
    category = str(feedback_data.get("import_category") or "").strip()
    category_key = GOOGLE_FORM_ENTRY_IDS["import_category"]
    if category in GOOGLE_FORM_IMPORT_CATEGORIES or not category:
        form_values[category_key] = category
    else:
        form_values[category_key] = "__other_option__"
        form_values[f"{category_key}.other_option_response"] = category

    request = urllib.request.Request(
        GOOGLE_FORM_RESPONSE_URL,
        data=urllib.parse.urlencode(form_values).encode("utf-8"),
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "User-Agent": "FXGuard-AI-Feedback/1.0",
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=8) as response:
        status_code = response.getcode()
    if 200 <= status_code < 400:
        return {"enabled": True, "status": "saved", "method": "google_form"}
    return {
        "enabled": True,
        "status": "failed",
        "method": "google_form",
        "detail": f"Google Forms returned HTTP {status_code}.",
    }


def load_model(currency: str):
    import joblib

    try:
        currency = validate_currency(currency)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    key = f"model_{currency}_flexible"
    if key not in _cache:
        path = MODEL_DIR / f"risk_model_{currency}_flexible.pkl"
        if not path.exists():
            raise HTTPException(status_code=500, detail=f"Model not found: {path}")
        _cache[key] = joblib.load(path)
    return _cache[key]


def load_features():
    if "features" not in _cache:
        import pandas as pd

        path = DATA_DIR / "exchange_rates_features_and_labels_4year.csv"
        if not path.exists():
            raise HTTPException(status_code=500, detail="Feature dataset is missing.")
        df = pd.read_csv(path, parse_dates=["date"])
        df = df.sort_values("date").reset_index(drop=True)
        _cache["features"] = df
    return _cache["features"]


def load_daily_calendar():
    if "daily" not in _cache:
        import pandas as pd

        path = DATA_DIR / "exchange_rates_daily_calendar_4year.csv"
        if not path.exists():
            raise HTTPException(status_code=500, detail="Daily calendar dataset is missing.")
        df = pd.read_csv(path, parse_dates=["date"])
        df = df.sort_values("date").reset_index(drop=True)
        _cache["daily"] = df
    return _cache["daily"]


def latest_feature_row():
    df = load_features()
    required = df.dropna(subset=FEATURE_COLUMNS)
    if required.empty:
        raise HTTPException(status_code=500, detail="No feature row available for prediction.")
    return required.iloc[-1]


def risk_considerations(risk: str, currency: str = "USD") -> List[str]:
    pair = f"{currency}/RWF"
    if risk == "High":
        return [
            "Ask your supplier whether you can split the payment or adjust the due date if that would ease cash-flow pressure.",
            "Set aside at least the displayed planning buffer in RWF before the payment is due.",
            f"Check the latest {pair} rate again before authorising payment, then confirm the final RWF amount with your bank or payment provider.",
        ]
    if risk == "Medium":
        return [
            "Include the displayed planning buffer in your cash-flow plan so a moderate rate change does not disrupt the payment.",
            "Confirm the invoice due date and ask whether a partial payment is acceptable if available cash is tight.",
            f"Check the latest {pair} rate again before paying and compare the updated RWF cost with today's estimate.",
        ]
    return [
        "Use today's estimated RWF cost as your starting point and keep the displayed planning buffer available for normal rate movement.",
        "Confirm the supplier's due date and payment details early to avoid a last-minute currency conversion.",
        f"Check the latest {pair} rate once more before paying, especially if the due date is several days away.",
    ]


def risk_pressure_rate(risk: str, horizon: int) -> float:
    """Scale the illustrative planning scenario with the payment period."""
    horizon_scale = (horizon / 7) ** 0.5
    base_rate = {"Low": 0.0025, "Medium": 0.006, "High": 0.012}[risk]
    return min(base_rate * horizon_scale, {"Low": 0.02, "Medium": 0.05, "High": 0.10}[risk])


def model_predict(currency: str, horizon: int):
    import pandas as pd

    model = load_model(currency)
    try:
        row = latest_currency_feature_row(currency, FEATURE_COLUMNS)
    except (RuntimeError, FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    model_values = row[FEATURE_COLUMNS].astype(float).to_dict()
    model_values["horizon_days"] = float(horizon)
    X = pd.DataFrame([model_values], columns=MODEL_FEATURE_COLUMNS)
    pred = str(model.predict(X)[0])

    confidence = None
    predicted_probability = None
    top_probability_label = None
    probabilities = {}
    if hasattr(model, "predict_proba"):
        try:
            classes = list(model.classes_)
            probs = model.predict_proba(X)[0]
            probabilities = {str(cls): round(float(prob), 4) for cls, prob in zip(classes, probs)}
            top_probability_index = int(max(range(len(probs)), key=lambda i: probs[i]))
            top_probability_label = str(classes[top_probability_index])
            predicted_probability = round(float(probs[top_probability_index]), 4)
            confidence = predicted_probability
        except Exception:
            probabilities = {}
            confidence = None
            predicted_probability = None
            top_probability_label = None
    else:
        confidence = 1.0
        predicted_probability = 1.0
        top_probability_label = pred

    return pred, confidence, predicted_probability, top_probability_label, probabilities, row


@app.get("/")
def home():
    index = FRONTEND_DIR / "index.html"
    if index.exists():
        return FileResponse(index)
    return {"message": "FXGuard AI API is running. Visit /docs for API documentation."}


@app.get("/health")
def health():
    return {
        "status": "ok",
        "project": "FXGuard AI",
        "currencies": list(SUPPORTED_CURRENCIES),
        "pairs": [f"{currency}/RWF" for currency in SUPPORTED_CURRENCIES],
        "rate_provider": provider_status(),
    }


@app.get("/api/currencies")
def currencies():
    return {
        "base_currency": "RWF",
        "currencies": [
            {"code": code, "pair": f"{code}/RWF", **details}
            for code, details in CURRENCY_INFO.items()
        ],
    }


@app.get("/api/latest-rate")
def latest_rate(currency: str = "USD"):
    try:
        code = validate_currency(currency)
        df = combined_daily(code)
    except (ValueError, FileNotFoundError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    row = df.iloc[-1]
    return {
        "date": str(row["date"].date()),
        "currency": code,
        "currency_name": CURRENCY_INFO[code]["name"],
        "currency_symbol": CURRENCY_INFO[code]["symbol"],
        "pair": f"{code}/RWF",
        "buying_rate": round(float(row["buying_rate"]), 4),
        "selling_rate": round(float(row["selling_rate"]), 4),
        "mid_rate": round(float(row["mid_rate"]), 4),
        "is_official_observation": bool(row.get("is_official_observation", 1)),
        "source": row.get("source", "National Bank of Rwanda Excel export"),
        "rate_type": row.get("rate_type", "BNR buying/average/selling rates"),
        "provider_status": provider_status(),
    }


@app.get("/api/latest-rates")
def latest_rates():
    return {
        "base_currency": "RWF",
        "rates": [latest_rate(currency) for currency in SUPPORTED_CURRENCIES],
    }


@app.get("/api/data-freshness")
def data_freshness(currency: str = "USD"):
    try:
        code = validate_currency(currency)
        daily = combined_daily(code)
    except (ValueError, FileNotFoundError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    latest_rate_date = daily["date"].max().date()
    latest_feature_date = latest_currency_feature_row(code, FEATURE_COLUMNS)["date"].date()
    today = datetime.now().date()
    days_since_rate = max((today - latest_rate_date).days, 0)
    days_since_features = max((today - latest_feature_date).days, 0)

    if days_since_rate <= 5:
        status = "fresh"
    elif days_since_rate <= 30:
        status = "aging"
    else:
        status = "stale"

    return {
        "status": status,
        "currency": code,
        "pair": f"{code}/RWF",
        "today": str(today),
        "latest_rate_date": str(latest_rate_date),
        "latest_feature_date": str(latest_feature_date),
        "days_since_latest_rate": days_since_rate,
        "days_since_latest_features": days_since_features,
        "source": str(daily.iloc[-1].get("source", "National Bank of Rwanda Excel export")),
        "provider_status": provider_status(),
        "note": "Rates come from the latest manually imported official BNR Excel exports; import newer exports to refresh them.",
    }


@app.get("/api/history")
def history(days: int = 90, currency: str = "USD"):
    try:
        code = validate_currency(currency)
        df = combined_daily(code).tail(max(7, min(days, 1461)))
    except (ValueError, FileNotFoundError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {
        "currency": code,
        "pair": f"{code}/RWF",
        "source": str(df.iloc[-1].get("source", "BNR reference-rate dataset")),
        "points": [
            {"date": str(r.date.date()), "mid_rate": round(float(r.mid_rate), 4)}
            for r in df.itertuples()
        ],
    }


@app.get("/api/model-metadata")
def model_metadata():
    return load_json(MODEL_DIR / "flexible_horizon_model_metadata.json")


@app.post("/api/predict-risk")
def predict_risk(req: RiskRequest):
    try:
        currency = validate_currency(req.currency)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    today = rwanda_today()
    horizon_days = (req.payment_date - today).days
    if not MIN_PAYMENT_DAYS <= horizon_days <= MAX_PAYMENT_DAYS:
        raise HTTPException(
            status_code=400,
            detail="Payment date must be between 1 and 100 days from today.",
        )

    risk, confidence, predicted_probability, top_probability_label, probabilities, row = model_predict(currency, horizon_days)
    current_rate = float(row["mid_rate"])
    current_cost = req.amount * current_rate
    pressure_rate = risk_pressure_rate(risk, horizon_days)
    possible_extra_cost = current_cost * pressure_rate
    planning_buffer = current_cost * max(pressure_rate, 0.0025)
    considerations = risk_considerations(risk, currency)

    metadata = load_json(MODEL_DIR / "flexible_horizon_model_metadata.json")
    model_evidence = metadata.get("models", {}).get(currency, {})
    reliability = model_evidence.get(
        "reliability",
        {"status": "not_evaluated", "gate_is_project_defined": True},
    )
    return {
        "currency": currency,
        "currency_name": CURRENCY_INFO[currency]["name"],
        "currency_symbol": CURRENCY_INFO[currency]["symbol"],
        "pair": f"{currency}/RWF",
        "payment_date": str(req.payment_date),
        "horizon_days": horizon_days,
        "days_to_payment": horizon_days,
        "amount": req.amount,
        "amount_currency": req.amount,
        "amount_usd": req.amount if currency == "USD" else None,
        "analysis_date": str(row["date"].date()),
        "model_version": metadata.get("generated_at", "2.0.0"),
        "current_rate": round(current_rate, 4),
        "current_cost_rwf": round(current_cost, 2),
        "risk_level": risk,
        "confidence": confidence,
        "confidence_score": confidence,
        "model_score": confidence,
        "predicted_probability": predicted_probability,
        "top_probability_label": top_probability_label,
        "class_probabilities": probabilities,
        "probability_distribution": probabilities,
        "score_is_approximate": True,
        "score_calibration": "uncalibrated",
        "model_reliability": reliability,
        "operational_use": (
            "experimental_decision_support"
            if reliability.get("status") != "meets_provisional_gate"
            else "passed_provisional_research_gate"
        ),
        "assumed_pressure_rate": pressure_rate,
        "possible_extra_cost_rwf": round(possible_extra_cost, 2),
        "planning_buffer_estimate_rwf": round(planning_buffer, 2),
        "key_drivers": {
            "daily_return": round(float(row["daily_return"]), 6),
            "return_7d": round(float(row["return_7d"]), 6),
            "return_14d": round(float(row["return_14d"]), 6),
            "ma_7": round(float(row["ma_7"]), 4),
            "ma_30": round(float(row["ma_30"]), 4),
            "ma_gap": round(float(row["ma_gap"]), 6),
            "volatility_7d": round(float(row["volatility_7d"]), 6),
            "momentum_7d": round(float(row["momentum_7d"]), 6),
            "spread_pct": round(float(row["spread_pct"]), 6),
            "depreciation_days_7d": int(row["depreciation_days_7d"]),
        },
        "considerations": considerations,
        "rate_source": str(row.get("source", "BNR reference-rate dataset")),
        "rate_type": str(row.get("rate_type", "BNR reference rate")),
        "disclaimer": "This is an early estimate for planning. It cannot promise the future exchange rate or remove currency risk. Do not make a payment decision from this result alone.",
    }


def build_excel_report(result: dict) -> BytesIO:
    # Spreadsheet support is export-only. Import it here so it does not slow
    # the API's cold start or the first dashboard request.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Risk Assessment"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 36
    sheet.column_dimensions["B"].width = 74

    title_fill = PatternFill("solid", fgColor="102A46")
    section_fill = PatternFill("solid", fgColor="F5CA52")
    header_fill = PatternFill("solid", fgColor="FFF2B6")

    sheet.append(["FXGUARD AI — PAYMENT CHECK REPORT"])
    sheet.merge_cells("A1:B1")
    sheet["A1"].fill = title_fill
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    def add_section(title: str, headings: tuple[str, str]) -> None:
        sheet.append([])
        sheet.append([title])
        row_number = sheet.max_row
        sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=2)
        sheet.cell(row_number, 1).fill = section_fill
        sheet.cell(row_number, 1).font = Font(color="102A46", bold=True)
        sheet.append(list(headings))
        for cell in sheet[sheet.max_row]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

    amount = result.get("amount", result.get("amount_currency"))
    add_section("SUMMARY", ("Field", "Value"))
    summary_rows = [
        ("Date created", datetime.now().strftime("%d %B %Y")),
        ("BNR rate date", result["analysis_date"]),
        ("Currency", result["pair"]),
        (f"Payment amount ({result['currency']})", amount),
        ("Payment date", result["payment_date"]),
        ("Days to payment", result["horizon_days"]),
        (f"Current rate (RWF per {result['currency']})", result["current_rate"]),
        ("Risk level", result["risk_level"]),
        ("How strongly FXGuard chose this level", result.get("confidence_score", result.get("confidence"))),
        ("Cost at current rate (RWF)", result["current_cost_rwf"]),
        ("Possible extra cost (RWF)", result["possible_extra_cost_rwf"]),
        ("Planning buffer estimate (RWF)", result["planning_buffer_estimate_rwf"]),
        ("Rate source", result["rate_source"]),
    ]
    for label, value in summary_rows:
        sheet.append([label, value])
    for row_number in range(7, sheet.max_row + 1):
        label = sheet.cell(row_number, 1).value
        value_cell = sheet.cell(row_number, 2)
        if label == "How strongly FXGuard chose this level":
            value_cell.number_format = '"≈ "0.0%'
        elif label and ("amount" in label.lower() or "rate" in label.lower() or "cost" in label.lower() or "buffer" in label.lower()):
            value_cell.number_format = "#,##0.00"

    add_section("HOW FXGUARD COMPARED THE RISK LEVELS", ("Risk level", "Approx. result"))
    for risk_class, probability in sorted(
        result.get("class_probabilities", {}).items(), key=lambda item: item[1], reverse=True
    ):
        sheet.append([risk_class, probability])
        sheet.cell(sheet.max_row, 2).number_format = '"≈ "0.0%'

    add_section("PAYMENT TIPS", ("#", "What you may consider"))
    for index, consideration in enumerate(result.get("considerations", []), start=1):
        sheet.append([index, consideration])
        sheet.cell(sheet.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    plain_driver_labels = {
        "daily_return": "Change since the previous day",
        "return_7d": "Change over the last 7 days",
        "return_14d": "Change over the last 14 days",
        "ma_7": "Typical rate over the last 7 days",
        "ma_30": "Typical rate over the last 30 days",
        "ma_gap": "Difference between recent and longer-term rates",
        "volatility_7d": "How much the rate moved up and down this week",
        "momentum_7d": "Direction the rate moved this week",
        "spread_pct": "Gap between BNR buying and selling rates",
        "depreciation_days_7d": "Days the foreign currency became more expensive this week",
    }
    percentage_drivers = {
        "daily_return", "return_7d", "return_14d", "ma_gap",
        "volatility_7d", "momentum_7d", "spread_pct",
    }

    add_section("WHAT THIS RESULT IS BASED ON", ("Recent rate information", "Value"))
    for signal, value in result.get("key_drivers", {}).items():
        if signal in percentage_drivers:
            display_value = f"{float(value) * 100:+.2f}%"
        elif signal in {"ma_7", "ma_30"}:
            display_value = f"{float(value):,.4f} RWF"
        elif signal == "depreciation_days_7d":
            display_value = f"{int(value)} of 7 days"
        else:
            display_value = value
        sheet.append([plain_driver_labels.get(signal, signal.replace("_", " ").title()), display_value])

    add_section("IMPORTANT NOTE", ("Notice", "Details"))
    sheet.append(["For planning only", result["disclaimer"]])
    sheet.cell(sheet.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@app.post("/api/export-excel")
def export_excel(req: RiskRequest):
    result = predict_risk(req)
    output = build_excel_report(result)
    filename = f"fxguard-result-{result['analysis_date']}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/feedback")
def feedback(req: FeedbackRequest):
    import pandas as pd

    FEEDBACK_DIR.mkdir(parents=True, exist_ok=True)
    feedback_data = req.model_dump()
    feedback_data["submitted_at"] = datetime.now().isoformat(timespec="seconds")
    row = pd.DataFrame([feedback_data], columns=FEEDBACK_COLUMNS)

    if FEEDBACK_FILE.exists():
        existing = pd.read_excel(FEEDBACK_FILE, dtype={"phone_number": str})
        combined = pd.concat([existing, row], ignore_index=True)
    else:
        combined = row

    combined = combined.reindex(columns=FEEDBACK_COLUMNS)
    combined.to_excel(FEEDBACK_FILE, index=False)
    try:
        google_sheet_result = append_feedback_to_google_sheet(feedback_data)
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
        google_sheet_result = {
            "enabled": True,
            "status": "failed",
            "detail": str(exc),
        }
    saved_remotely = google_sheet_result.get("status") == "saved"
    return {
        "status": "saved" if saved_remotely else "saved_locally",
        "message": (
            "Thank you for your feedback."
            if saved_remotely
            else "Your feedback was saved locally, but the response sheet could not be reached."
        ),
        "file": project_path_label(FEEDBACK_FILE),
        "google_sheet": google_sheet_result,
    }


@app.get("/api/feedback-file")
def download_feedback_file():
    if not FEEDBACK_FILE.exists():
        raise HTTPException(status_code=404, detail="No feedback responses have been saved yet.")
    return FileResponse(
        FEEDBACK_FILE,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=FEEDBACK_FILE.name,
    )
